"""Generate a mechanical broker-data reconciliation and trading audit.

This command is read-only with respect to the database.  It writes only the
requested Markdown report.  All currency labels use ``Rs`` for cp1252 console
compatibility.
"""
from __future__ import annotations

import argparse
import html
import math
import re
import sqlite3
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any, Iterable, Sequence

from manas_os.engine import eod_detectors
from manas_os.tools.import_broker import Fill, Match, aggregate_fills, fifo_match, read_tradebooks


def _norm(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ").replace("%", " pct ")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if math.isnan(float(value)) else float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--", "NA", "N/A"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.+-]", "", text.strip("()"))
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return None


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _dates_in_value(value: object) -> list[date]:
    direct = _as_date(value)
    if direct:
        return [direct]
    text = str(value or "")
    found: list[date] = []
    patterns = (
        (r"\b\d{4}-\d{2}-\d{2}\b", "%Y-%m-%d"),
        (r"\b\d{1,2}[-/]\d{1,2}[-/]\d{4}\b", None),
        (r"\b\d{1,2}[- ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]\d{4}\b", None),
    )
    for pattern, fixed_format in patterns:
        for token in re.findall(pattern, text, flags=re.IGNORECASE):
            if fixed_format:
                try:
                    found.append(datetime.strptime(token, fixed_format).date())
                except ValueError:
                    pass
            else:
                parsed = _as_date(token)
                if parsed:
                    found.append(parsed)
    return found


def _fmt_money(value: float | None) -> str:
    return "no data" if value is None else f"Rs {value:,.2f}"


def _fmt_pct(value: float | None) -> str:
    return "no data" if value is None else f"{value:,.2f}%"


def _md(value: object) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")


class Report:
    def __init__(self) -> None:
        self.lines: list[str] = []

    def heading(self, title: str, level: int = 2) -> None:
        self.lines.extend(("", f"{'#' * level} {title}", ""))

    def text(self, value: str = "") -> None:
        self.lines.append(value)

    def table(self, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
        material = list(rows)
        self.lines.append("| " + " | ".join(_md(x) for x in headers) + " |")
        self.lines.append("| " + " | ".join("---" for _ in headers) + " |")
        if material:
            self.lines.extend("| " + " | ".join(_md(x) for x in row) + " |" for row in material)
        else:
            self.lines.append("| " + " | ".join(["no data"] + [""] * (len(headers) - 1)) + " |")

    def render(self) -> str:
        return "\n".join(self.lines).lstrip() + "\n"


@dataclass(frozen=True)
class SheetTable:
    source: str
    sheet: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, object], ...]
    metadata_dates: tuple[date, ...] = ()


@dataclass
class ReportMetrics:
    source: str
    pnl: float | None
    date_min: date | None
    date_max: date | None
    charges: dict[str, float]
    warnings: list[str]


def _matrix_to_tables(source: Path, sheet: str, matrix: Sequence[Sequence[object]]) -> list[SheetTable]:
    """Find offset tables headed by Symbol, or charge-oriented tax tables."""
    tables: list[SheetTable] = []
    starts: list[int] = []
    for idx, row in enumerate(matrix):
        names = {_norm(cell) for cell in row if cell is not None}
        vertical_charges = "amount" in names and bool(
            names & {"particular", "particulars", "description", "charge", "charges"}
        )
        symbol_header = bool(names & {"symbol", "instrument", "tradingsymbol", "security_symbol", "scrip_name"})
        horizontal_charge_headers = sum(
            1 for name in names
            if "charge" in name or name in {"brokerage", "stt", "stt_ctt", "gst", "stamp_duty", "sebi_fees"}
        ) >= 2
        if symbol_header or vertical_charges or horizontal_charge_headers:
            starts.append(idx)
    for pos, start in enumerate(starts):
        raw_headers = list(matrix[start])
        last = max((idx for idx, value in enumerate(raw_headers) if value is not None), default=-1)
        if last < 0:
            continue
        seen: dict[str, int] = defaultdict(int)
        headers = []
        for idx, value in enumerate(raw_headers[: last + 1]):
            base = _norm(value) or f"column_{idx + 1}"
            seen[base] += 1
            headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        end = starts[pos + 1] if pos + 1 < len(starts) else len(matrix)
        rows: list[dict[str, object]] = []
        blank_run = 0
        for raw in matrix[start + 1 : end]:
            values = list(raw[: len(headers)]) + [None] * max(0, len(headers) - len(raw))
            if not any(value not in (None, "") for value in values):
                blank_run += 1
                if blank_run >= 2:
                    break
                continue
            blank_run = 0
            rows.append(dict(zip(headers, values)))
        if rows:
            tables.append(SheetTable(str(source), sheet, tuple(headers), tuple(rows)))
    return tables


def read_xlsx_tables(path: str | Path) -> list[SheetTable]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx broker reports") from exc
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        tables = []
        for sheet in workbook.worksheets:
            matrix = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            tables.extend(_matrix_to_tables(source, sheet.title, matrix))
            metadata_dates = tuple(sorted({
                parsed for row in matrix for value in row
                for parsed in _dates_in_value(value)
            }))
            if metadata_dates:
                tables.append(SheetTable(str(source), sheet.title, (), (), metadata_dates))
        return tables
    finally:
        workbook.close()


def read_legacy_tables(path: str | Path) -> tuple[list[SheetTable], list[str]]:
    """Read real BIFF .xls or HTML-disguised .xls; skip honestly on failure."""
    source = Path(path)
    warnings: list[str] = []
    prefix = source.read_bytes()[:512].lower()
    try:
        import pandas as pd
        if b"<html" in prefix or b"<table" in prefix or b"<!doctype" in prefix:
            frames = pd.read_html(source)
        else:
            frames = pd.read_excel(source, sheet_name=None, engine="xlrd")
            frames = list(frames.values())
        tables = []
        for idx, frame in enumerate(frames):
            matrix = [tuple(frame.columns)] + list(frame.itertuples(index=False, name=None))
            tables.extend(_matrix_to_tables(source, f"table_{idx + 1}", matrix))
        return tables, warnings
    except Exception as exc:  # optional legacy dependency/file corruption boundary
        warnings.append(f"{source}: unparseable legacy .xls; skipped ({type(exc).__name__}: {exc})")
        return [], warnings


def read_report_tables(path: str | Path) -> tuple[list[SheetTable], list[str]]:
    source = Path(path)
    if source.suffix.lower() == ".xls":
        return read_legacy_tables(source)
    try:
        return read_xlsx_tables(source), []
    except Exception as exc:
        return [], [f"{source}: report skipped ({type(exc).__name__}: {exc})"]


PNL_PRIORITY = (
    "realized_p_and_l", "realised_p_and_l", "realized_pnl", "realised_pnl",
    "net_realized_p_and_l", "net_realised_p_and_l", "net_p_and_l", "net_pnl",
    "p_and_l", "pnl", "profit_loss", "profit",
)
CHARGE_TERMS = (
    "charge", "brokerage", "stt", "transaction_charge", "exchange_transaction", "gst", "stamp",
    "sebi", "dp_charge", "clearing_charge", "auction_penalty", "ipft",
)


def report_metrics(path: str | Path, charges_only: bool = False) -> ReportMetrics:
    tables, warnings = read_report_tables(path)
    pnl_total = 0.0
    pnl_found = False
    dates: list[date] = []
    charges: dict[str, float] = defaultdict(float)
    for table in tables:
        dates.extend(table.metadata_dates)
        pnl_column = next((name for name in PNL_PRIORITY if name in table.headers), None)
        charge_columns = [
            name for name in table.headers
            if any(term in name for term in CHARGE_TERMS) and "total" not in name
        ]
        label_column = next((name for name in ("charge", "charges", "particular", "particulars", "description") if name in table.headers), None)
        amount_column = next((name for name in ("amount", "value", "total") if name in table.headers), None)
        for row in table.rows:
            label = " ".join(str(value or "") for value in row.values()).strip().lower()
            is_total = label.startswith("total ") or label == "total" or "grand total" in label
            for name, value in row.items():
                if "date" in name:
                    parsed = _as_date(value)
                    if parsed:
                        dates.append(parsed)
            if pnl_column and not charges_only and not is_total:
                value = _number(row.get(pnl_column))
                if value is not None:
                    pnl_total += value
                    pnl_found = True
            if not is_total:
                for column in charge_columns:
                    value = _number(row.get(column))
                    if value is not None:
                        charges[column] += value
                if label_column and amount_column:
                    label_name = _norm(row.get(label_column))
                    value = _number(row.get(amount_column))
                    if label_name and value is not None and any(term in label_name for term in CHARGE_TERMS):
                        charges[label_name] += value
    return ReportMetrics(
        str(path), pnl_total if pnl_found else None,
        min(dates) if dates else None, max(dates) if dates else None,
        dict(sorted(charges.items())), warnings,
    )


@dataclass
class Cycle:
    symbol: str
    cycle: int
    direction: str
    entry_date: date
    exit_date: date
    entry_at: datetime
    exit_at: datetime
    qty: float
    entry_notional: float
    exit_notional: float
    pnl: float

    @property
    def return_pct(self) -> float:
        return self.pnl / self.entry_notional * 100.0 if self.entry_notional else 0.0

    @property
    def holding_days(self) -> int:
        return (self.exit_date - self.entry_date).days


def build_cycles(matches: Sequence[Match]) -> list[Cycle]:
    groups: dict[tuple[str, int], list[Match]] = defaultdict(list)
    for match in matches:
        groups[(match.symbol, match.cycle)].append(match)
    cycles = []
    for (symbol, cycle_no), parts in sorted(groups.items(), key=lambda item: min(x.entry_at for x in item[1])):
        entry_notional = sum(part.entry_price * part.qty for part in parts)
        exit_notional = sum(part.exit_price * part.qty for part in parts)
        cycles.append(Cycle(
            symbol, cycle_no, parts[0].direction,
            min(part.entry_at for part in parts).date(), max(part.exit_at for part in parts).date(),
            min(part.entry_at for part in parts), max(part.exit_at for part in parts),
            sum(part.qty for part in parts), entry_notional, exit_notional,
            sum(part.pnl for part in parts),
        ))
    return cycles


def fiscal_year(day: date) -> int:
    """Return the opening calendar year of the Indian Apr-Mar fiscal year."""
    return day.year if day.month >= 4 else day.year - 1


def _period_label(opening_year: int) -> str:
    return f"FY{opening_year}-{str(opening_year + 1)[-2:]}"


def _bucket(value: float, breaks: Sequence[tuple[float, str]]) -> str:
    for upper, label in breaks:
        if value <= upper:
            return label
    return breaks[-1][1]


def _group_stats(cycles: Sequence[Cycle], key_fn) -> list[tuple[str, int, str, str]]:
    groups: dict[str, list[Cycle]] = defaultdict(list)
    for cycle in cycles:
        groups[key_fn(cycle)].append(cycle)
    rows = []
    for key, values in groups.items():
        rows.append((key, len(values), _fmt_pct(sum(x.pnl > 0 for x in values) / len(values) * 100.0),
                     _fmt_money(mean(x.pnl for x in values))))
    return rows


def _cycle_summary(cycles: Sequence[Cycle], charges: float) -> list[tuple[str, str]]:
    wins = [cycle for cycle in cycles if cycle.pnl > 0]
    losses = [cycle for cycle in cycles if cycle.pnl < 0]
    avg_win_pct = mean(c.return_pct for c in wins) if wins else None
    avg_loss_pct = mean(c.return_pct for c in losses) if losses else None
    payoff = (avg_win_pct / abs(avg_loss_pct)) if avg_win_pct is not None and avg_loss_pct not in (None, 0) else None
    gross = sum(c.pnl for c in cycles)
    turnover = sum(c.entry_notional + c.exit_notional for c in cycles)
    return [
        ("Total trades (FIFO round trips)", str(len(cycles))),
        ("Win rate", _fmt_pct(len(wins) / len(cycles) * 100.0) if cycles else "no data"),
        ("Average winner", _fmt_pct(avg_win_pct)),
        ("Average loser", _fmt_pct(avg_loss_pct)),
        ("Payoff (avg win % / abs avg loss %)", "no data" if payoff is None else f"{payoff:.2f}"),
        ("Expectancy / trade", _fmt_money(gross / len(cycles)) if cycles else "no data"),
        ("Gross realized P&L", _fmt_money(gross)),
        ("Charges", _fmt_money(charges)),
        ("Net of charges", _fmt_money(gross - charges)),
        ("Turnover", _fmt_money(turnover)),
        ("Charges / turnover", _fmt_pct(charges / turnover * 100.0) if turnover else "no data"),
    ]


def _render_period(report: Report, title: str, cycles: Sequence[Cycle], charges: float) -> None:
    report.heading(title)
    if cycles:
        report.text(f"Observed tradebook range: {min(c.entry_date for c in cycles)} through {max(c.exit_date for c in cycles)}.")
    else:
        report.text("Observed tradebook range: no data.")
    report.table(("Metric", "Value"), _cycle_summary(cycles, charges))

    report.heading("Monthly P&L", 3)
    monthly: dict[str, float] = defaultdict(float)
    for cycle in cycles:
        monthly[cycle.exit_date.strftime("%Y-%m")] += cycle.pnl
    report.table(("Month", "Realized P&L"), ((month, _fmt_money(pnl)) for month, pnl in sorted(monthly.items())))

    report.heading("Hold-time buckets", 3)
    hold_breaks = ((0, "same-day"), (2, "1-2d"), (5, "3-5d"), (14, "1-2w"), (10**9, ">2w"))
    report.table(("Bucket", "Trades", "Win rate", "Avg P&L"),
                 _group_stats(cycles, lambda c: _bucket(c.holding_days, hold_breaks)))

    report.heading("Position-size buckets", 3)
    size_breaks = ((4999.999, "< Rs 5000"), (9999.999, "Rs 5000-9999"),
                   (24999.999, "Rs 10000-24999"), (49999.999, "Rs 25000-49999"),
                   (10**30, "Rs 50000+"))
    report.table(("Bucket", "Trades", "Win rate", "Avg P&L"),
                 _group_stats(cycles, lambda c: _bucket(c.entry_notional, size_breaks)))

    by_symbol: dict[str, float] = defaultdict(float)
    for cycle in cycles:
        by_symbol[cycle.symbol] += cycle.pnl
    report.heading("Top-10 winner symbols", 3)
    report.table(("Symbol", "P&L"), ((s, _fmt_money(v)) for s, v in sorted(by_symbol.items(), key=lambda x: x[1], reverse=True)[:10]))
    report.heading("Top-10 loser symbols", 3)
    report.table(("Symbol", "P&L"), ((s, _fmt_money(v)) for s, v in sorted(by_symbol.items(), key=lambda x: x[1])[:10] if v < 0))

    report.heading("Repeat-cycle symbols", 3)
    repeated: dict[str, list[Cycle]] = defaultdict(list)
    for cycle in cycles:
        repeated[cycle.symbol].append(cycle)
    report.table(
        ("Symbol", "Cycle", "Entry", "Exit", "P&L"),
        ((symbol, idx, cycle.entry_date, cycle.exit_date, _fmt_money(cycle.pnl))
         for symbol, values in sorted(repeated.items()) if len(values) >= 3
         for idx, cycle in enumerate(sorted(values, key=lambda c: c.entry_at), 1)),
    )


def _charge_total(metrics: Sequence[ReportMetrics]) -> float:
    return sum(sum(item.charges.values()) for item in metrics)


def _metric_period(item: ReportMetrics) -> int | None:
    # Report start date identifies the covered FY.  A later workbook-generation
    # timestamp must not move a prior-FY report into the current FY.
    anchor = item.date_min or item.date_max
    return fiscal_year(anchor) if anchor else None


def _behavior_flags(report: Report, cycles: Sequence[Cycle], fills: Sequence[Fill]) -> None:
    report.heading("BEHAVIOR FLAGS")
    report.heading("Same-day round trips", 3)
    same_day = [c for c in cycles if c.holding_days == 0]
    report.table(("Symbol", "Date", "Entry size", "P&L"),
                 ((c.symbol, c.entry_date, _fmt_money(c.entry_notional), _fmt_money(c.pnl)) for c in same_day))

    report.heading("Averaging into intraday fallers", 3)
    open_intraday_buys: dict[tuple[str, date], list[list[float]]] = defaultdict(list)
    averaging = []
    for fill in sorted(fills, key=lambda f: f.executed_at):
        key = (fill.symbol, fill.trade_date)
        lots = open_intraday_buys[key]
        if fill.side == "SELL":
            remaining = fill.qty
            while remaining > 1e-9 and lots:
                closed = min(remaining, lots[0][0])
                lots[0][0] -= closed
                remaining -= closed
                if lots[0][0] <= 1e-9:
                    lots.pop(0)
            continue
        cheaper_than = [lot for lot in lots if fill.price < lot[1]]
        if cheaper_than:
            prior = cheaper_than[-1]
            drop = (fill.price - prior[1]) / prior[1] * 100.0
            averaging.append((fill.symbol, fill.executed_at, fill.qty, fill.price, prior[1], f"{drop:.2f}%"))
        lots.append([fill.qty, fill.price])
    report.table(("Symbol", "Later buy", "Qty", "Later price", "Prior buy", "Price change"), averaging)

    report.heading("Open-chase buys 09:15-09:45", 3)
    chase = [fill for fill in fills if fill.side == "BUY" and (9, 15) <= (fill.executed_at.hour, fill.executed_at.minute) <= (9, 45)]
    report.table(("Symbol", "Timestamp", "Qty", "Price", "Notional"),
                 ((f.symbol, f.executed_at, f.qty, f"{f.price:.2f}", _fmt_money(f.qty * f.price)) for f in chase))

    report.heading("Re-entry within 5 observed sessions of a losing exit", 3)
    sessions = sorted({fill.trade_date for fill in fills})
    session_index = {day: idx for idx, day in enumerate(sessions)}
    entries: dict[str, list[Cycle]] = defaultdict(list)
    for cycle in cycles:
        entries[cycle.symbol].append(cycle)
    reentries = []
    for symbol, values in entries.items():
        ordered = sorted(values, key=lambda c: c.entry_at)
        for loser in (cycle for cycle in ordered if cycle.pnl < 0):
            later = next((cycle for cycle in ordered if cycle.entry_at > loser.exit_at), None)
            if later and loser.exit_date in session_index and later.entry_date in session_index:
                gap = session_index[later.entry_date] - session_index[loser.exit_date]
                if 0 <= gap <= 5:
                    reentries.append((symbol, loser.exit_date, _fmt_money(loser.pnl), later.entry_date, gap, _fmt_money(later.entry_notional)))
    report.table(("Symbol", "Losing exit", "Loss", "Re-entry", "Observed sessions", "Re-entry size"), reentries)

    report.heading("Disposition effect", 3)
    winners = [c.holding_days for c in cycles if c.pnl > 0]
    losers = [c.holding_days for c in cycles if c.pnl < 0]
    report.table(("Cohort", "Trades", "Average hold days"), (
        ("Winners", len(winners), f"{mean(winners):.2f}" if winners else "no data"),
        ("Losers", len(losers), f"{mean(losers):.2f}" if losers else "no data"),
    ))
    if winners and losers:
        verdict = "losers held longer" if mean(losers) > mean(winners) else "winners held at least as long"
        report.text(f"Mechanical comparison: {verdict}; difference {abs(mean(losers) - mean(winners)):.2f} days.")

    report.heading("Position sizes below Rs 5000", 3)
    tiny = [c for c in cycles if c.entry_notional < 5000]
    report.table(("Symbol", "Entry", "Size", "Exit", "P&L"),
                 ((c.symbol, c.entry_date, _fmt_money(c.entry_notional), c.exit_date, _fmt_money(c.pnl)) for c in tiny))


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    return conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone() is not None


def _tool_lens(report: Report, conn: sqlite3.Connection, cycles: Sequence[Cycle]) -> None:
    report.heading("TOOL LENS")
    report.text("2026-period entry-date lookup only; this section reads persisted rows and does not run a scan.")
    required = ("scan_candidates", "discovery_bucket", "refusals")
    missing = [name for name in required if not _table_exists(conn, name)]
    if missing:
        report.text("No data: missing database tables " + ", ".join(missing) + ".")
        return
    rows = []
    counts = defaultdict(int)
    seen = set()
    for cycle in sorted(cycles, key=lambda c: c.entry_at):
        key = (cycle.entry_date.isoformat(), cycle.symbol)
        if key in seen:
            continue
        seen.add(key)
        scan = conn.execute("SELECT 1 FROM scan_candidates WHERE scan_date=? AND symbol=? LIMIT 1", key).fetchone() is not None
        discovery = conn.execute("SELECT 1 FROM discovery_bucket WHERE scan_date=? AND symbol=? LIMIT 1", key).fetchone() is not None
        refusal = conn.execute("SELECT 1 FROM refusals WHERE scan_date=? AND symbol=? LIMIT 1", key).fetchone() is not None
        status = "surfaced" if scan or discovery else "refused" if refusal else "never-saw"
        counts[status] += 1
        rows.append((key[0], key[1], "yes" if scan else "no", "yes" if discovery else "no", "yes" if refusal else "no", status))
    report.table(("Entry date", "Symbol", "scan_candidates", "discovery_bucket", "refusals", "Classification"), rows)
    report.table(("Classification", "Entry-date x symbol count"),
                 ((status, counts[status]) for status in ("surfaced", "refused", "never-saw")))


SYMBOL_COLUMNS = ("symbol", "instrument", "tradingsymbol", "security_symbol", "scrip_name", "security_name")
QTY_COLUMNS = ("quantity", "qty", "net_qty", "net_quantity", "total_quantity")
COST_COLUMNS = ("average_price", "avg_price", "average_cost", "avg_cost", "buy_avg", "cost_price")


def holdings_from_file(path: str | Path) -> tuple[dict[str, tuple[float, float]], list[str]]:
    tables, warnings = read_report_tables(path)
    holdings: dict[str, tuple[float, float]] = {}
    for table in tables:
        symbol_col = next((name for name in SYMBOL_COLUMNS if name in table.headers), None)
        qty_col = next((name for name in QTY_COLUMNS if name in table.headers), None)
        cost_col = next((name for name in COST_COLUMNS if name in table.headers), None)
        if not (symbol_col and qty_col and cost_col):
            continue
        for row in table.rows:
            symbol = str(row.get(symbol_col) or "").strip().upper()
            qty, cost = _number(row.get(qty_col)), _number(row.get(cost_col))
            if symbol and symbol not in {"TOTAL", "GRAND TOTAL"} and qty is not None and cost is not None:
                holdings[symbol] = (qty, cost)
    if not holdings:
        warnings.append(f"{path}: no holdings table with symbol, quantity, and average cost was found")
    return holdings, warnings


def _load_open_lots(conn: sqlite3.Connection) -> dict[str, tuple[float, float]]:
    if not _table_exists(conn, "broker_open_lots"):
        return {}
    return {row[0]: (float(row[1]), float(row[2])) for row in conn.execute(
        "SELECT symbol, qty, avg_cost FROM broker_open_lots"
    )}


def _bars(conn: sqlite3.Connection, symbol: str) -> list[dict[str, Any]]:
    if not _table_exists(conn, "daily_prices"):
        return []
    rows = conn.execute(
        "SELECT trade_date, open, high, low, close, prev_close, volume "
        "FROM daily_prices WHERE symbol=? AND series='EQ' ORDER BY trade_date DESC LIMIT 260",
        (symbol,),
    ).fetchall()
    return [dict(row) for row in reversed(rows)]


def _holdings_exit(report: Report, conn: sqlite3.Connection,
                   holdings_files: Sequence[str], warnings: list[str]) -> None:
    report.heading("HOLDINGS X EXIT ENGINE")
    file_holdings: dict[str, tuple[float, float]] = {}
    if holdings_files:
        file_holdings, file_warnings = holdings_from_file(holdings_files[0])
        warnings.extend(file_warnings)
        for supplemental in holdings_files[1:]:
            _, extra_warnings = holdings_from_file(supplemental)
            warnings.extend(extra_warnings)
    fifo_holdings = _load_open_lots(conn)
    report.text("The first --holdings file is authoritative; later files are parsed only for legacy-read diagnostics.")
    report.table(("Symbol", "File qty", "FIFO qty", "Qty gap", "File avg", "FIFO avg"), (
        (symbol,
         file_holdings.get(symbol, (None, None))[0] if symbol in file_holdings else "no data",
         fifo_holdings.get(symbol, (None, None))[0] if symbol in fifo_holdings else "no data",
         f"{file_holdings[symbol][0] - fifo_holdings[symbol][0]:.4f}" if symbol in file_holdings and symbol in fifo_holdings else "no data",
         f"{file_holdings[symbol][1]:.2f}" if symbol in file_holdings else "no data",
         f"{fifo_holdings[symbol][1]:.2f}" if symbol in fifo_holdings else "no data")
        for symbol in sorted(set(file_holdings) | set(fifo_holdings))
    ))
    engine_rows = []
    for symbol in sorted(set(file_holdings) | set(fifo_holdings)):
        qty, avg_cost = file_holdings[symbol] if symbol in file_holdings else fifo_holdings[symbol]
        bars = _bars(conn, symbol)
        if not bars or bars[-1].get("close") is None:
            engine_rows.append((symbol, qty, f"{avg_cost:.2f}", "no data", "no data", "no data", "no data", "no data"))
            continue
        last = float(bars[-1]["close"])
        state = eod_detectors.exit_state(bars)
        closes = [float(bar["close"]) if bar.get("close") is not None else None for bar in bars]
        sma50 = eod_detectors.sma(closes, 50)[-1]
        sma200 = eod_detectors.sma(closes, 200)[-1]
        relation = ", ".join((
            f"50SMA {'above' if last >= sma50 else 'below'} ({sma50:.2f})" if sma50 is not None else "50SMA no data",
            f"200SMA {'above' if last >= sma200 else 'below'} ({sma200:.2f})" if sma200 is not None else "200SMA no data",
        ))
        rules = "; ".join(rule.get("rule", "") for rule in state.get("fired_rules", [])) or "none"
        pnl_pct = (last - avg_cost) / avg_cost * 100.0 if avg_cost else None
        engine_rows.append((symbol, f"{qty:.4f}", f"{avg_cost:.2f}", f"{last:.2f}", _fmt_pct(pnl_pct),
                            state.get("state", "no data"), rules, relation))
    report.table(("Symbol", "Qty", "Avg cost", "Last close", "P&L%", "Exit state", "Fired rules", "Close vs 50/200SMA"), engine_rows)
    report.text("Mechanical engine classifications only.")


def build_audit(db_path: str | Path, tradebooks: Sequence[str], pnl_paths: Sequence[str],
                taxpnl_path: str | None, holdings_paths: Sequence[str]) -> str:
    raw_rows, duplicate_count = read_tradebooks(tradebooks)
    fills = aggregate_fills(raw_rows)
    matches, _ = fifo_match(fills)
    cycles = build_cycles(matches)
    pnl_metrics = [report_metrics(path) for path in pnl_paths]
    tax_metrics = [report_metrics(taxpnl_path)] if taxpnl_path else []
    warnings = [warning for item in pnl_metrics + tax_metrics for warning in item.warnings]
    by_fy: dict[int, list[Cycle]] = defaultdict(list)
    for cycle in cycles:
        by_fy[fiscal_year(cycle.exit_date)].append(cycle)
    fiscal_years = sorted(by_fy)
    current_fy = fiscal_years[-1] if fiscal_years else None
    previous_fy = fiscal_years[-2] if len(fiscal_years) >= 2 else None

    report = Report()
    report.text("# Broker Import Audit")
    report.text("")
    report.text(f"Tradebook rows after trade_id dedupe: {len(raw_rows)}; duplicates removed: {duplicate_count}.")
    report.heading("RECONCILIATION")
    reconciliation_rows = []
    for fy in fiscal_years:
        fifo_pnl = sum(c.pnl for c in by_fy[fy])
        applicable = [item for item in pnl_metrics if _metric_period(item) == fy]
        if fy == previous_fy:
            applicable += tax_metrics
        period_charges = _charge_total(tax_metrics) if fy == previous_fy else 0.0
        if not applicable:
            reconciliation_rows.append((_period_label(fy), "no dated broker report", _fmt_money(fifo_pnl),
                                        "no data", "no data", _fmt_money(period_charges),
                                        _fmt_money(fifo_pnl - period_charges)))
        for item in applicable:
            gap = fifo_pnl - item.pnl if item.pnl is not None else None
            reconciliation_rows.append((_period_label(fy), Path(item.source).name, _fmt_money(fifo_pnl),
                                        _fmt_money(item.pnl), _fmt_money(gap), _fmt_money(period_charges),
                                        _fmt_money(fifo_pnl - period_charges)))
    report.table(("Period", "Broker source", "FIFO realized", "Broker reported", "FIFO - broker gap", "Tax charges", "FIFO net of charges"), reconciliation_rows)
    report.text("Broker workbooks without parseable dates remain unassigned below; their totals are not forced into a period.")
    report.table(("Source", "Date range", "Reported P&L"), (
        (item.source, f"{item.date_min or 'unknown'} to {item.date_max or 'unknown'}", _fmt_money(item.pnl))
        for item in pnl_metrics + tax_metrics
    ))
    charge_rows = []
    for item in tax_metrics:
        charge_rows.extend((item.source, name, _fmt_money(value)) for name, value in item.charges.items())
    report.heading("Tax P&L charges", 3)
    report.table(("Source", "Charge column", "Total"), charge_rows)

    previous_cycles = by_fy.get(previous_fy, []) if previous_fy is not None else []
    current_cycles = by_fy.get(current_fy, []) if current_fy is not None else []
    _render_period(report, "FY2025-26 (last year, observed data)", previous_cycles,
                   _charge_total(tax_metrics) if previous_fy is not None else 0.0)
    _render_period(report, "APR-JUL 2026 (observed current period)", current_cycles, 0.0)
    _behavior_flags(report, cycles, fills)

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        _tool_lens(report, conn, current_cycles)
        _holdings_exit(report, conn, holdings_paths, warnings)
    finally:
        conn.close()

    report.heading("DATA QUALITY / SKIPS")
    if warnings:
        for warning in warnings:
            report.text(f"- {html.escape(warning)}")
    else:
        report.text("No parser skips recorded.")
    report.text("- FIFO uses tradebook executions only; taxes and charges are not allocated to individual round trips.")
    report.text("- Observed-session re-entry counts use sessions present in the supplied tradebooks, not a full exchange calendar.")
    return report.render()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Audit Zerodha broker exports against Manas data")
    parser.add_argument("--db", required=True)
    parser.add_argument("--tradebooks", nargs="+", required=True, help="All tradebook CSV files")
    parser.add_argument("--pnl", nargs="*", default=[], help="P&L .xlsx/.xls files")
    parser.add_argument("--taxpnl", help="Tax P&L .xlsx file")
    parser.add_argument("--holdings", nargs="+", required=True, help="Current holdings first; optional legacy files after")
    parser.add_argument("--out", required=True, help="Markdown output path")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        content = build_audit(args.db, args.tradebooks, args.pnl, args.taxpnl, args.holdings)
        output = Path(args.out)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(content, encoding="utf-8")
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(f"wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
